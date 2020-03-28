import openpyxl

class excelUtility(object):

    @staticmethod
    def getNumberOfRows(filePath, sheetName):
        wb = openpyxl.load_workbook(filename=filePath)
        sheet = wb[sheetName]
        return sheet.max_row

    @staticmethod
    def getNumberOfColumns(filePath, sheetName):
        wb = openpyxl.load_workbook(filename=filePath)
        sheet = wb[sheetName]
        return sheet.max_column

    @staticmethod
    def readDataFromExcel(filePath, sheetName, xRow, yCol):
        wb = openpyxl.load_workbook(filename=filePath)
        sheet = wb[sheetName]
        return sheet.cell(row=xRow, column=yCol).value

    @staticmethod
    def writeDataToExcel(filePath, sheetName, xRow, yCol, data):
        wb = openpyxl.load_workbook(filename=filePath)
        sheet = wb[sheetName]
        sheet.cell(row=xRow,column=yCol).value = data
        wb.save(filename=filePath)